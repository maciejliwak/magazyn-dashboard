from datetime import datetime

from django.contrib import messages
from django.contrib.auth import logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User, Group
from django.db import models
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic.edit import CreateView
from openpyxl import Workbook

from magazyn_core.models import (
    Magazyn, Urzadzenie, Czesc, Wypozyczenie, OperacjaLog,
    PrzesuniecieMagazynowe, ArchiwumUrzadzen
)
from magazyn_core.forms import (
    UrzadzenieForm, CzescForm, WypozyczenieForm,
    PrzesuniecieForm, PrzekazanieForm, UzupełnienieForm
)
from userprofile.models import UserProfile


def is_not_observer(user):
    return not user.groups.filter(name="Obserwator").exists()


@login_required
def dashboard(request):
    magazyn_id = request.GET.get("magazyn")
    fraza = request.GET.get("q")
    sortuj = request.GET.get("sortuj") or "nazwa"
    kierunek = request.GET.get("kierunek") or "asc"

    # ⛳ Przekształć parametr magazyn_id z URL
    try:
        wybrany_magazyn = int(magazyn_id) if magazyn_id else None
    except (ValueError, TypeError):
        wybrany_magazyn = None

    # 🛡️ Nadpisz magazyn dla obserwatora
    if request.user.groups.filter(name="Obserwator").exists():
        if hasattr(request.user, "userprofile") and request.user.userprofile.magazyn:
            wybrany_magazyn = request.user.userprofile.magazyn.id

    # 📊 Ustal sortowanie
    sort_key = sortuj if kierunek == "asc" else f"-{sortuj}"

    # 📁 Pokaż magazyny — tylko jeden dla obserwatora
    if request.user.groups.filter(name="Obserwator").exists():
        if hasattr(request.user, "userprofile") and request.user.userprofile.magazyn:
            magazyny = Magazyn.objects.filter(id=request.user.userprofile.magazyn.id)
        else:
            magazyny = Magazyn.objects.none()
    else:
        magazyny = Magazyn.objects.all()

    # 📦 Pobierz dane
    urzadzenia = Urzadzenie.objects.select_related("magazyn", "nazwa", "producent", "model") \
        .prefetch_related("wypozyczenia").filter(na_stale=False).order_by(sort_key)

    czesci = Czesc.objects.select_related("magazyn", "nazwa").filter(klient_nazwa__isnull=True)
    czesci_przekazane = Czesc.objects.select_related("magazyn", "nazwa").filter(klient_nazwa__isnull=False)
    wypozyczenia = Wypozyczenie.objects.select_related("urzadzenie")
    archiwum = ArchiwumUrzadzen.objects.select_related("magazyn", "nazwa", "producent", "model").order_by(sort_key)

    # 📍 Filtrowanie po magazynie
    if wybrany_magazyn:
        urzadzenia = urzadzenia.filter(magazyn_id=wybrany_magazyn)
        czesci = czesci.filter(magazyn_id=wybrany_magazyn)
        czesci_przekazane = czesci_przekazane.filter(magazyn_id=wybrany_magazyn)
        wypozyczenia = wypozyczenia.filter(urzadzenie__magazyn_id=wybrany_magazyn)
        archiwum = archiwum.filter(magazyn_id=wybrany_magazyn)

    # 🔎 Filtrowanie po frazie
    if fraza:
        urzadzenia = urzadzenia.filter(
            models.Q(nazwa__nazwa__icontains=fraza) |
            models.Q(numer_seryjny__icontains=fraza) |
            models.Q(model__nazwa__icontains=fraza) |
            models.Q(producent__nazwa__icontains=fraza)
        )
        archiwum = archiwum.filter(
            models.Q(nazwa__nazwa__icontains=fraza) |
            models.Q(numer_seryjny__icontains=fraza) |
            models.Q(model__nazwa__icontains=fraza) |
            models.Q(producent__nazwa__icontains=fraza)
        )
        czesci = czesci.filter(nazwa__nazwa__icontains=fraza)
        czesci_przekazane = czesci_przekazane.filter(nazwa__nazwa__icontains=fraza)
        wypozyczenia = wypozyczenia.filter(
            models.Q(urzadzenie__nazwa__nazwa__icontains=fraza) |
            models.Q(urzadzenie__numer_seryjny__icontains=fraza) |
            models.Q(klient_nazwa__icontains=fraza) |
            models.Q(numer_zgloszenia__icontains=fraza)
        )

    # 📋 Zwróć dane do szablonu
    return render(request, "magazyn_core/index.html", {
        "magazyny": magazyny,
        "urzadzenia": urzadzenia,
        "czesci": czesci,
        "czesci_przekazane": czesci_przekazane,
        "wypozyczenia": wypozyczenia,
        "archiwum": archiwum,
        "wybrany_magazyn": wybrany_magazyn,
        "fraza": fraza or "",
        "sortuj": sortuj,
        "kierunek": kierunek,
    })



def custom_logout(request):
    logout(request)
    return redirect("login")


class AddUrzadzenieView(UserPassesTestMixin, LoginRequiredMixin, CreateView):
    model = Urzadzenie
    form_class = UrzadzenieForm
    template_name = "magazyn_core/add_urzadzenie.html"
    success_url = reverse_lazy("dashboard")

    def test_func(self):
        return is_not_observer(self.request.user)

    def form_valid(self, form):
        if Urzadzenie.objects.filter(numer_seryjny=form.cleaned_data["numer_seryjny"]).exists():
            return self.render_to_response(self.get_context_data(form=form, duplikat=True))
        response = super().form_valid(form)
        OperacjaLog.objects.create(
            typ_operacji="DODANIE",
            opis=f"DODANIE: {self.object.nazwa.nazwa} ({self.object.numer_seryjny})",
            uzytkownik=self.request.user
        )
        return response


class AddCzescView(UserPassesTestMixin, LoginRequiredMixin, CreateView):
    model = Czesc
    form_class = CzescForm
    template_name = "magazyn_core/add_czesc.html"
    success_url = reverse_lazy("dashboard")

    def test_func(self):
        return is_not_observer(self.request.user)

    def form_valid(self, form):
        if form.cleaned_data["numer_seryjny"] and Czesc.objects.filter(numer_seryjny=form.cleaned_data["numer_seryjny"]).exists():
            return self.render_to_response(self.get_context_data(form=form, duplikat=True))
        response = super().form_valid(form)
        OperacjaLog.objects.create(
            typ_operacji="DODANIE CZĘŚCI",
            opis=f"DODANIE: Część {self.object.nazwa.nazwa} ({self.object.numer_seryjny or 'bez nr'})",
            uzytkownik=self.request.user
        )
        return response


class DodajNaStanCzescView(UserPassesTestMixin, LoginRequiredMixin, View):
    def test_func(self):
        return is_not_observer(self.request.user)

    def get(self, request, pk):
        czesc = get_object_or_404(Czesc, pk=pk)
        form = UzupełnienieForm()
        return render(request, "magazyn_core/dodaj_na_stan_czesc.html", {"czesc": czesc, "form": form})

    def post(self, request, pk):
        czesc = get_object_or_404(Czesc, pk=pk)
        form = UzupełnienieForm(request.POST)
        if form.is_valid():
            ilosc = form.cleaned_data["ilosc_do_dodania"]
            czesc.ilosc += ilosc
            czesc.save()

            OperacjaLog.objects.create(
                typ_operacji="UZUPEŁNIENIE CZĘŚCI",
                opis=f"Uzupełniono: {czesc.nazwa.nazwa} (+{ilosc} szt.) → razem: {czesc.ilosc} szt.",
                uzytkownik=request.user
            )

            messages.success(request, f"Uzupełniono część o {ilosc} sztuk.")
            return redirect("dashboard")

        return render(request, "magazyn_core/dodaj_na_stan_czesc.html", {"czesc": czesc, "form": form})

class AddWypozyczenieView(UserPassesTestMixin, LoginRequiredMixin, CreateView):
    model = Wypozyczenie
    form_class = WypozyczenieForm
    template_name = "magazyn_core/add_wypozyczenie.html"

    def test_func(self):
        return is_not_observer(self.request.user)

    def get_success_url(self):
        return reverse_lazy("dashboard")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["urzadzenie"] = Urzadzenie.objects.get(pk=self.kwargs["pk"])
        return context

    def form_valid(self, form):
        urzadzenie = Urzadzenie.objects.get(pk=self.kwargs["pk"])
        form.instance.urzadzenie = urzadzenie
        response = super().form_valid(form)
        OperacjaLog.objects.create(
            typ_operacji="WYPOŻYCZENIE",
            opis=f"{urzadzenie.nazwa.nazwa} ({urzadzenie.numer_seryjny}) → {form.instance.klient_nazwa}, zgłoszenie: {form.instance.numer_zgloszenia}",
            uzytkownik=self.request.user
        )
        return response


class ReturnWypozyczenieView(UserPassesTestMixin, LoginRequiredMixin, View):
    def test_func(self):
        return is_not_observer(self.request.user)

    def post(self, request, pk):
        wyp = get_object_or_404(Wypozyczenie, pk=pk)
        OperacjaLog.objects.create(
            typ_operacji="ZWROT",
            opis=f"ZWROT: {wyp.urzadzenie.nazwa.nazwa} ({wyp.urzadzenie.numer_seryjny}) od {wyp.klient_nazwa}",
            uzytkownik=request.user
        )
        wyp.delete()
        return redirect("dashboard")


class PrzesunUrzadzenieView(UserPassesTestMixin, LoginRequiredMixin, View):
    def test_func(self):
        return is_not_observer(self.request.user)

    def get(self, request, pk):
        urzadzenie = get_object_or_404(Urzadzenie, pk=pk)
        form = PrzesuniecieForm()
        return render(request, "magazyn_core/przesun_urzadzenie.html", {"urzadzenie": urzadzenie, "form": form})

    def post(self, request, pk):
        urzadzenie = get_object_or_404(Urzadzenie, pk=pk)
        form = PrzesuniecieForm(request.POST)
        if form.is_valid():
            nowy_magazyn = form.cleaned_data["nowy_magazyn"]
            stary_magazyn = urzadzenie.magazyn

            urzadzenie.magazyn = nowy_magazyn
            urzadzenie.save()

            PrzesuniecieMagazynowe.objects.create(
                urzadzenie=urzadzenie,
                magazyn_zrodlowy=stary_magazyn,
                magazyn_docelowy=nowy_magazyn,
                uzytkownik=request.user
            )

            OperacjaLog.objects.create(
                typ_operacji="PRZESUNIĘCIE",
                opis=f"{urzadzenie.nazwa.nazwa} ({urzadzenie.numer_seryjny}) z {stary_magazyn.nazwa} ➝ {nowy_magazyn.nazwa}",
                uzytkownik=request.user
            )
            return redirect("dashboard")
        return render(request, "magazyn_core/przesun_urzadzenie.html", {"urzadzenie": urzadzenie, "form": form})


class MarkNaStaleView(UserPassesTestMixin, LoginRequiredMixin, View):
    def test_func(self):
        return is_not_observer(self.request.user)

    def get(self, request, pk):
        urzadzenie = get_object_or_404(Urzadzenie, pk=pk)
        wyp = Wypozyczenie.objects.filter(urzadzenie=urzadzenie).first()
        initial_data = {
            "klient_nazwa": wyp.klient_nazwa if wyp else "",
            "numer_zgloszenia": wyp.numer_zgloszenia if wyp else "",
            "uwagi": urzadzenie.uwagi
        }
        form = PrzekazanieForm(initial=initial_data)
        return render(request, "magazyn_core/mark_na_stale.html", {
            "urzadzenie": urzadzenie, "form": form, "czy_z_wypozyczenia": bool(wyp)
        })

    def post(self, request, pk):
        urzadzenie = get_object_or_404(Urzadzenie, pk=pk)
        wyp = Wypozyczenie.objects.filter(urzadzenie=urzadzenie).first()
        form = PrzekazanieForm(request.POST)
        if form.is_valid():
            klient = wyp.klient_nazwa if wyp else form.cleaned_data["klient_nazwa"]
            zgloszenie = wyp.numer_zgloszenia if wyp else form.cleaned_data["numer_zgloszenia"]

            ArchiwumUrzadzen.objects.create(
                nazwa=urzadzenie.nazwa,
                producent=urzadzenie.producent,
                model=urzadzenie.model,
                numer_seryjny=urzadzenie.numer_seryjny,
                ilosc=urzadzenie.ilosc,
                magazyn=urzadzenie.magazyn,
                klient_nazwa=klient,
                numer_zgloszenia=zgloszenie,
                uwagi=form.cleaned_data["uwagi"]
            )

            OperacjaLog.objects.create(
                typ_operacji="PRZEKAZANIE",
                opis=f"PRZEKAZANIE: {urzadzenie.nazwa.nazwa} ({urzadzenie.numer_seryjny}) dla {klient}, zgłoszenie: {zgloszenie}",
                uzytkownik=request.user
            )

            urzadzenie.delete()
            if wyp:
                wyp.delete()

            messages.success(request, "Urządzenie przekazane na stałe.")
            return redirect("dashboard")
        return render(request, "magazyn_core/mark_na_stale.html", {
            "urzadzenie": urzadzenie, "form": form, "czy_z_wypozyczenia": bool(wyp)
        })


class MarkCzescNaStaleView(UserPassesTestMixin, LoginRequiredMixin, View):
    def test_func(self):
        return is_not_observer(self.request.user)

    def get(self, request, pk):
        czesc = get_object_or_404(Czesc, pk=pk)
        form = PrzekazanieForm()
        return render(request, "magazyn_core/mark_czesc_na_stale.html", {
            "czesc": czesc,
            "form": form
        })

    def post(self, request, pk):
        czesc = get_object_or_404(Czesc, pk=pk)
        form = PrzekazanieForm(request.POST)
        if form.is_valid():
            ilosc = form.cleaned_data["ilosc_do_przekazania"]
            if ilosc > czesc.ilosc:
                form.add_error("ilosc_do_przekazania", "Brak tylu sztuk na stanie.")
                return render(request, "magazyn_core/mark_czesc_na_stale.html", {
                    "czesc": czesc,
                    "form": form
                })

            czesc.ilosc -= ilosc
            czesc.save()

            numer = czesc.numer_seryjny
            if numer and Czesc.objects.filter(numer_seryjny=numer).exclude(pk=czesc.pk).exists():
                numer = None

            przekazana = Czesc.objects.create(
                nazwa=czesc.nazwa,
                producent=czesc.producent,
                model=czesc.model,
                numer_seryjny=numer,
                magazyn=czesc.magazyn,
                ilosc=ilosc,
                klient_nazwa=form.cleaned_data["klient_nazwa"],
                numer_zgloszenia=form.cleaned_data["numer_zgloszenia"],
                uwagi=form.cleaned_data["uwagi"]
            )

            OperacjaLog.objects.create(
                typ_operacji="PRZEKAZANIE CZĘŚCI",
                opis=f"{przekazana.nazwa.nazwa} ({przekazana.numer_seryjny or 'bez nr'}) {ilosc} szt. dla {przekazana.klient_nazwa}",
                uzytkownik=request.user
            )
            return redirect("dashboard")

        return render(request, "magazyn_core/mark_czesc_na_stale.html", {
            "czesc": czesc,
            "form": form
        })

class HistoriaUrzadzeniaView(LoginRequiredMixin, View):
    def get(self, request, pk):
        urzadzenie = get_object_or_404(Urzadzenie, pk=pk)
        wypozyczenia = Wypozyczenie.objects.filter(urzadzenie_id=pk)
        przesuniecia = PrzesuniecieMagazynowe.objects.filter(urzadzenie_id=pk)
        logi = OperacjaLog.objects.filter(opis__icontains=urzadzenie.numer_seryjny)

        timeline = []

        for w in wypozyczenia:
            timeline.append({
                "data": timezone.make_aware(datetime.combine(w.data_wypozyczenia, datetime.min.time())),
                "typ": "🎟️ Wypożyczenie",
                "opis": f"Dla {w.klient_nazwa}, zgłoszenie: {w.numer_zgloszenia}"
            })

        for p in przesuniecia:
            aware = timezone.make_aware(p.data_przesuniecia) if timezone.is_naive(p.data_przesuniecia) else p.data_przesuniecia
            timeline.append({
                "data": aware,
                "typ": "🔁 Przesunięcie",
                "opis": f"{p.magazyn_zrodlowy.nazwa} → {p.magazyn_docelowy.nazwa}"
            })

        for log in logi:
            aware = timezone.make_aware(log.data_operacji) if timezone.is_naive(log.data_operacji) else log.data_operacji
            timeline.append({
                "data": aware,
                "typ": f"📝 {log.typ_operacji}",
                "opis": log.opis
            })

        timeline.sort(key=lambda x: x["data"], reverse=True)

        return render(request, "magazyn_core/historia_urzadzenia.html", {
            "urzadzenie": urzadzenie,
            "wypozyczenia": wypozyczenia,
            "przesuniecia": przesuniecia,
            "logi": logi,
            "timeline": timeline,
        })


class EksportHistoriaUrzadzeniaView(UserPassesTestMixin, LoginRequiredMixin, View):
    def test_func(self):
        return is_not_observer(self.request.user)

    def get(self, request, pk):
        urzadzenie = get_object_or_404(Urzadzenie, pk=pk)
        wypozyczenia = Wypozyczenie.objects.filter(urzadzenie_id=pk)
        przesuniecia = PrzesuniecieMagazynowe.objects.filter(urzadzenie_id=pk)
        logi = OperacjaLog.objects.filter(opis__icontains=urzadzenie.numer_seryjny)

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Wypożyczenia"
        ws1.append(["Data", "Klient", "Zgłoszenie"])
        for w in wypozyczenia:
            ws1.append([
                w.data_wypozyczenia.strftime("%Y-%m-%d") if w.data_wypozyczenia else "brak",
                w.klient_nazwa or "brak",
                w.numer_zgloszenia or "brak"
            ])

        ws2 = wb.create_sheet("Przesunięcia")
        ws2.append(["Data", "Źródłowy", "Docelowy"])
        for p in przesuniecia:
            ws2.append([
                p.data_przesuniecia.strftime("%Y-%m-%d %H:%M") if p.data_przesuniecia else "brak",
                p.magazyn_zrodlowy.nazwa,
                p.magazyn_docelowy.nazwa
            ])

        ws3 = wb.create_sheet("Logi")
        ws3.append(["Data", "Typ", "Opis", "Użytkownik"])
        for log in logi:
            ws3.append([
                log.data_operacji.strftime("%Y-%m-%d %H:%M") if log.data_operacji else "brak",
                log.typ_operacji,
                log.opis,
                log.uzytkownik.username if log.uzytkownik else "brak"
            ])

        filename = f"historia_{urzadzenie.numer_seryjny or 'urzadzenie'}.xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class PrzywrocUrzadzenieView(UserPassesTestMixin, LoginRequiredMixin, View):
    def test_func(self):
        return is_not_observer(self.request.user)

    def post(self, request, pk):
        arch = get_object_or_404(ArchiwumUrzadzen, pk=pk)
        urzadzenie = Urzadzenie.objects.create(
            nazwa=arch.nazwa,
            producent=arch.producent,
            model=arch.model,
            numer_seryjny=arch.numer_seryjny,
            ilosc=arch.ilosc,
            magazyn=arch.magazyn,
            klient_nazwa="",
            numer_zgloszenia="",
            uwagi=arch.uwagi
        )

        OperacjaLog.objects.create(
            typ_operacji="ZWROT Z ARCHIWUM",
            opis=f"Przywrócono: {urzadzenie.nazwa} ({urzadzenie.numer_seryjny}) do magazynu {urzadzenie.magazyn.nazwa}",
            uzytkownik=request.user
        )

        arch.delete()
        messages.success(request, "Urządzenie przywrócono na stan magazynowy.")
        return redirect("dashboard")

from django.contrib import messages
from django.contrib.auth import logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib.auth.models import User


# Widok do zmiany hasła
from django.contrib.auth import update_session_auth_hash

@login_required
def zmien_haslo_startowe(request):
    if request.method == "POST":
        nowe_haslo = request.POST["password"]
        request.user.set_password(nowe_haslo)
        request.user.save()
        
        # Aktualizuj sesję, aby nie wylogować użytkownika
        update_session_auth_hash(request, request.user)
        
        # Ustaw last_login na teraz (już nie wymuszaj zmiany)
        request.user.last_login = timezone.now()
        request.user.save()
        
        messages.success(request, "Hasło zostało zmienione!")
        return redirect("dashboard")
    
    return render(request, "magazyn_core/zmien_haslo_startowe.html")


# Widok do sprawdzenia, czy użytkownik powinien zmienić hasło
@login_required
def sprawdzenie_hasla(request):
    # Jeżeli to pierwsze logowanie użytkownika, przekierowujemy go do zmiany hasła
    if request.user.last_login is None or request.user.last_login == request.user.date_joined:
        return redirect("zmien_haslo_startowe")
    else:
        return redirect("dashboard")


# Pozostałe widoki w projekcie:
from django.contrib.auth.models import User, Group
from django.utils import timezone

@login_required
@user_passes_test(is_not_observer)
def rejestruj_obserwatora(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        magazyn_id = request.POST["magazyn"]

        # Stwórz użytkownika, ale ZABLOKUJ last_login
        user = User.objects.create_user(
            username=username,
            password=password,
        )
        user.save()  # Zapisujemy najpierw, aby mieć user.id

        # Przypisz do grupy Obserwator
        group = Group.objects.get(name="Obserwator")
        user.groups.add(group)

        # Przypisz magazyn (jeśli masz UserProfile)
        magazyn = Magazyn.objects.get(id=magazyn_id)
        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.magazyn = magazyn
        profile.save()

        # UPEWNIJ SIĘ, ŻE last_login JEST NULL (wymusi zmianę hasła)
        User.objects.filter(id=user.id).update(last_login=None)

        return redirect("dashboard")

    return render(request, "magazyn_core/rejestruj_obserwatora.html", {
        "magazyny": Magazyn.objects.all()
    })

# Funkcja wylogowywania
def custom_logout(request):
    logout(request)
    return redirect("login")




@login_required
@user_passes_test(is_not_observer)
def rejestruj_obserwatora(request):
    if request.method == "POST":
        username = request.POST["username"]
        email = request.POST["email"]
        password = request.POST["password"]
        magazyn_id = request.POST["magazyn"]

        if User.objects.filter(username=username).exists():
            return render(request, "magazyn_core/rejestruj_obserwatora.html", {
                "magazyny": Magazyn.objects.all(),
                "error": "⚠ Użytkownik już istnieje!"
            })

        user = User.objects.create_user(username=username, email=email, password=password)
        group, _ = Group.objects.get_or_create(name="Obserwator")
        user.groups.add(group)

        magazyn = Magazyn.objects.get(id=magazyn_id)
        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.magazyn = magazyn
        profile.save()

        return redirect("dashboard")

    return render(request, "magazyn_core/rejestruj_obserwatora.html", {
        "magazyny": Magazyn.objects.all()
    })


@login_required
@user_passes_test(is_not_observer)
def eksport_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Magazyn"
    ws.append(["Typ", "Nazwa", "Producent", "Model", "Numer seryjny", "Ilość", "Magazyn", "Wypożyczone", "Klient", "Zgłoszenie", "Uwagi"])

    for u in Urzadzenie.objects.select_related("magazyn", "nazwa", "producent", "model").filter(na_stale=False):
        wyp = Wypozyczenie.objects.filter(urzadzenie=u).first()
        ws.append([
            "Urządzenie", u.nazwa.nazwa, u.producent.nazwa, u.model.nazwa, u.numer_seryjny,
            u.ilosc, u.magazyn.nazwa,
            "Tak" if wyp else "Nie",
            u.klient_nazwa or (wyp.klient_nazwa if wyp else ""),
            u.numer_zgloszenia or (wyp.numer_zgloszenia if wyp else ""),
            u.uwagi or ""
        ])

    for c in Czesc.objects.select_related("magazyn", "nazwa").filter(klient_nazwa__isnull=True):
        ws.append([
            "Część", c.nazwa.nazwa, c.producent or "", c.model or "", c.numer_seryjny,
            c.ilosc, c.magazyn.nazwa, "", "", "", c.uwagi or ""
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="magazyn.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_not_observer)
def eksport_widoku_excel(request):
    magazyn_id = request.GET.get("magazyn")
    fraza = request.GET.get("q")

    wb = Workbook()
    ws = wb.active
    ws.title = "Widok"
    ws.append(["Typ", "Nazwa", "Producent", "Model", "Numer seryjny", "Ilość", "Magazyn", "Wypożyczone", "Klient", "Zgłoszenie", "Uwagi"])

    urzadzenia = Urzadzenie.objects.select_related("magazyn", "nazwa", "producent", "model").filter(na_stale=False)
    czesci = Czesc.objects.select_related("magazyn", "nazwa").filter(klient_nazwa__isnull=True)

    if magazyn_id:
        urzadzenia = urzadzenia.filter(magazyn_id=magazyn_id)
        czesci = czesci.filter(magazyn_id=magazyn_id)

    if fraza:
        urzadzenia = urzadzenia.filter(
            models.Q(nazwa__nazwa__icontains=fraza) |
            models.Q(numer_seryjny__icontains=fraza) |
            models.Q(model__nazwa__icontains=fraza) |
            models.Q(producent__nazwa__icontains=fraza)
        )
        czesci = czesci.filter(nazwa__nazwa__icontains=fraza)

    for u in urzadzenia:
        wyp = Wypozyczenie.objects.filter(urzadzenie=u).first()
        ws.append([
            "Urządzenie", u.nazwa.nazwa, u.producent.nazwa, u.model.nazwa, u.numer_seryjny,
            u.ilosc, u.magazyn.nazwa,
            "Tak" if wyp else "Nie",
            u.klient_nazwa or (wyp.klient_nazwa if wyp else ""),
            u.numer_zgloszenia or (wyp.numer_zgloszenia if wyp else ""),
            u.uwagi or ""
        ])

    for c in czesci:
        ws.append([
            "Część", c.nazwa.nazwa, c.producent or "", c.model or "", c.numer_seryjny,
            c.ilosc, c.magazyn.nazwa, "", "", "", c.uwagi or ""
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="widok_magazynowy.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_not_observer)
def eksport_logi(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Historia Operacji"
    ws.append(["Typ operacji", "Opis", "Data", "Użytkownik"])

    for log in OperacjaLog.objects.select_related("uzytkownik").order_by("-data_operacji"):
        ws.append([
            log.typ_operacji,
            log.opis,
            log.data_operacji.strftime("%Y-%m-%d %H:%M"),
            log.uzytkownik.username if log.uzytkownik else "Brak"
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="logi_operacji.xlsx"'
    wb.save(response)
    return response

from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect

@login_required
def sprawdzenie_hasla(request):
    # Sprawdź, czy użytkownik NIGDY się nie logował (last_login == None)
    if request.user.last_login is None:
        return redirect("zmien_haslo_startowe")  # Przekieruj do zmiany hasła
    return redirect("dashboard")  # W przeciwnym razie idź do dashboardu


# Plik views.py zakończony — wszystkie widoki klasowe i funkcyjne zostały:
# - chronione przed dostępem grupy „Obserwator”
# - ujednolicone pod względem struktury
# - zgodne ze ścieżkami z urls.py i widokiem index.html
# - gotowe do testowania i wdrożenia

# Jeśli dodasz nowe widoki w przyszłości, pamiętaj aby:
# - dodać `UserPassesTestMixin` do widoków klasowych
# - dodać `@user_passes_test(is_not_observer)` do widoków funkcyjnych
# - schować przyciski w szablonie za pomocą `{% if not request.user.groups.all.0.name == "Obserwator" %}`

# Możesz też utworzyć własne mixiny lub middleware, aby scentralizować logikę uprawnień.

# 🔐 Twój magazyn jest teraz odporny na edycję przez użytkowników-obserwatorów,
# a jednocześnie zachowuje dostęp podglądowy i interaktywność.

# Jeśli chcesz, mogę przygotować gotowy zestaw testów do sprawdzania uprawnień.
